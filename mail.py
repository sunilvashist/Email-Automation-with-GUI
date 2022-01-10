import smtplib
from csv import reader
from email.message import EmailMessage
import datetime
import xlrd
import openpyxl



class Email:

    def __init__(self,sender,subject,body,password,filename,user):
        self.subject = subject
        self.body    = body
        self.password  = password
        self.filename  = filename
        self.sender    = sender
        self.user      = user


    def sendEmail(self,client, smtp):

        try:

            email = EmailMessage()
            email['from'] = self.user
            # email['to'] = 'vashist1508@gmail.com'
            email['subject'] = self.subject
            email.set_content(self.body)
            email['to'] = client
            smtp.send_message(email)
        except Exception as E:
            print("Error in sending email => ",E)


    def readfile(self):

        emailList = []
        try:
            if self.filename.endswith('.xls'):
                wb = xlrd.open_workbook(self.filename)
                sheet = wb.sheet_by_index(0)
                sheet.cell_value(0, 0)

                for i in range(sheet.nrows):
                    emailList.append(sheet.cell_value(i, 0))

            elif self.filename.endswith('.csv'):
                with open(self.filename, 'r') as read_obj:
                    csv_reader = reader(read_obj)
                    header = next(csv_reader)
                    # Check file as empty

                    if header is not None:
                        # print(header)
                        emailList.append(header[0])
                        for row in csv_reader:
                            # print(row)
                            emailList.append(row[0])

            elif self.filename.endswith('.xlsx'):
                wb_obj = openpyxl.load_workbook(self.filename)

                sheet_obj = wb_obj.active
                m_row = sheet_obj.max_row

                # Loop will print all values
                # of first column
                for i in range(1, m_row + 1):
                    cell_obj = sheet_obj.cell(row=i, column=1)
                    emailList.append(cell_obj.value)

        except Exception as E:
            print("Error in file reading => ", E)

        return emailList


    def read_and_send(self):
        emails = self.readfile()

        t1 = datetime.datetime.now()
        try:
            with smtplib.SMTP(host='smtp.gmail.com', port=587) as smtp:
                smtp.ehlo()
                smtp.starttls()
                smtp.login(self.sender, self.password)
                for i in emails:
                    self.sendEmail(i, smtp)
        except:
            print("Error occured at read_send. ")
        t2 = datetime.datetime.now()

        print((t2 - t1))

